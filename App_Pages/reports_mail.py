import os
import io
import streamlit as st
import pandas as pd
import plotly.express as px
from Core.DB_Connection import get_connection
from Core.functions import insert_log
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import win32com.client
import tempfile
from streamlit_extras.stylable_container import stylable_container

# Function to fetch report data
@st.cache_data(ttl=300)
def fetch_report_data():
    query = """
    SELECT
        cd.insuranceNum AS 'Insurance Number',
        cd.mailNum AS 'Mail Number',
        cd.mailDate AS 'Mail Date',
        cd.route AS 'Route',
        iata.IATA AS 'Complain Location',
        iata.Country AS 'Complain Country',
        cd.Sector,
        cd.fltNum AS 'Flight Number',
        cd.fltDate AS 'Flight Date',
        com.[type] AS Commodity,
        cd.weight_Kg AS 'Weight KG',
        cd.totalParcelsNum AS 'Total Parcels',
        cd.damages_lossesParcelsNum As 'Damaged Parcels',
        ct.complainDescription AS 'Complain Type',
        cd.[claim_other] AS 'Other Complain',
        cd.claimantName AS 'Claimant Name',
        cd.claimantType AS 'Claimant Type',
        cd.complainReceivedDate AS 'Claim Date',
        cd.insuranceSentDate AS 'Insurance Sent',
        cd.compansationRequested AS 'Compensation Requested',
        c.Alpha_Code AS 'Req. Currency',
        cd.ExchangeEGP_requested AS 'Requested in EGP',
        cd.ExchangeUSD_requested AS 'Requested in USD',
        cd.compansationRequestedDate AS 'Req. Date',
        cd.compansationAccepted AS 'Compensation Accepted',
        CC.Alpha_Code AS 'Acc. Currency',
        cd.ExchangeEGP_accepted AS 'Accepted in EGP',
        cd.ExchangeUSD_accepted AS 'Accepted in USD',
        cd.compansationAcceptedDate AS 'Acc. Date',
        st.[status] AS 'Status',
        cd.status_other AS 'Other Status',
        cd.Notes
    FROM ComplainsDetails_mail cd
    LEFT JOIN IATACodes iata ON iata.iataID = cd.complainLocationID
    LEFT JOIN CommodityType com ON com.id = cd.CommodityID
    LEFT JOIN ComplainsType ct ON ct.compliantypeID = cd.complainTypeID
    LEFT JOIN Status st ON st.statusID = cd.statusID
    LEFT JOIN Currency c ON c.currencyId = cd.compansationRequestedCurrency
    LEFT JOIN Currency CC ON CC.currencyId = cd.compansationAcceptedCurrency
    """

    engine = get_connection()
    df = pd.read_sql(query, engine)


    # Data Types (Mail DATE only)
    df["Mail Date"] = pd.to_datetime(df["Mail Date"], errors="coerce").dt.date
    df["Weight KG"] = pd.to_numeric(df["Weight KG"], errors="coerce")
    df["Damaged Parcels"] = pd.to_numeric(df["Damaged Parcels"], errors="coerce")

    df["Requested in USD"] = pd.to_numeric(df["Requested in USD"], errors="coerce")

    # Extract Year from Insurance Number
    df["Year"] = (
        df["Insurance Number"]
        .astype(str)
        .str.split("/")
        .str[-1]
        .str.extract(r"(\d{2,4})")[0]
    )

    return df


# Main Reports Page
def reports_page_mail():

    if "excel_file_mail" not in st.session_state:
        st.session_state.excel_file_mail = None

    if "pdf_file_mail" not in st.session_state:
        st.session_state.pdf_file_mail = None

    st.title("Mail Claim Dashboard")

    df = fetch_report_data()

    # Sidebar Filters
    st.sidebar.header("Filters / Options")

    filter_cols = {
        "Year": "Year",
        "Complain Country": "Complain Country",
        "Complain Location": "Complain Location",
        "Sector": "Sector",
        "Commodity": "Commodity",
        "Claimant Type": "Claimant Type",
        "Status": "Status",
        "Complain Type": "Complain Type"
    }

    filters = {}

    # Date filter (DATE not datetime)
    df["Insurance Sent"] = pd.to_datetime(df["Insurance Sent"], errors="coerce")

    valid_dates = df["Insurance Sent"].dropna()

    if not valid_dates.empty:
        min_date = valid_dates.min().date()
        max_date = valid_dates.max().date()
    else:
        today = datetime.today()
        min_date = today
        max_date = today

    date_range = st.sidebar.date_input(
        "Insurance Sent Date Range",
        value=(min_date, max_date)
    )

    for label, col in filter_cols.items():
        options = sorted(df[col].dropna().unique())
        state_key = f"filter_{col}"

        # Initialize state ONCE
        if state_key not in st.session_state:
            st.session_state[state_key] = options.copy()

        # Layout: Expander | Button
        c1, c2 = st.sidebar.columns([5, 1])

        with c1:
            with st.expander(label, expanded=False):
                selected = st.multiselect(
                    f"Select {label}",
                    options,
                    key=state_key
                )
                filters[col] = selected

        with c2:
            st.button(
                "⟳",
                key=f"all_{col}",
                on_click=lambda k=state_key, o=options: st.session_state.update({k: o}),
                help=f"Select all {label}"
            )


    # USD Range Filter (Sidebar)
    # Ensure numeric
    df["Requested in USD"] = pd.to_numeric(df["Requested in USD"], errors="coerce")

    # Safe max
    max_usd = df["Requested in USD"].max()
    if pd.isna(max_usd):
        max_usd = 10000
    else:
        max_usd = int(max_usd)

    st.sidebar.markdown("### Requested USD Range")

    # ---- Initialize state ----
    if "usd_min" not in st.session_state:
        st.session_state.usd_min = 0
    if "usd_max" not in st.session_state:
        st.session_state.usd_max = min(5000, max_usd)
    if "usd_slider" not in st.session_state:
        st.session_state.usd_slider = (st.session_state.usd_min, st.session_state.usd_max)


    # ---- Callbacks ----
    def update_slider():
        st.session_state.usd_slider = (
            st.session_state.usd_min,
            st.session_state.usd_max
        )

    def update_inputs():
        st.session_state.usd_min, st.session_state.usd_max = st.session_state.usd_slider


    # ---- Number inputs (typing) ----
    col1, col2 = st.sidebar.columns(2)

    with col1:
        st.number_input(
            "Min USD",
            min_value=0,
            max_value=max_usd,
            key="usd_min",
            on_change=update_slider
        )

    with col2:
        st.number_input(
            "Max USD",
            min_value=0,
            max_value=max_usd,
            key="usd_max",
            on_change=update_slider
        )

    if st.session_state.usd_min > st.session_state.usd_max:
        st.sidebar.warning("Min USD cannot be greater than Max USD")

    # ---- Slider ----
    st.sidebar.slider(
        "Adjust Range",
        min_value=0,
        max_value=max_usd,
        key="usd_slider",
        on_change=update_inputs
    )

    # Apply Filters
    filtered_df = df.copy()

    for col, selected in filters.items():
        filtered_df = filtered_df[filtered_df[col].isin(selected)]


    # APPLY DATE FILTER
    if date_range and len(date_range) == 2:
        start_date, end_date = date_range

        if start_date and end_date:
            filtered_df = filtered_df[
                (pd.to_datetime(filtered_df["Insurance Sent"], errors="coerce").dt.date >= start_date) &
                (pd.to_datetime(filtered_df["Insurance Sent"], errors="coerce").dt.date <= end_date)
            ]
            
    # Apply USD Range Filter
    filtered_df = filtered_df[
        (filtered_df["Requested in USD"].fillna(0) >= st.session_state.usd_min) &
        (filtered_df["Requested in USD"].fillna(0) <= st.session_state.usd_max)
    ]

    st.success(f"Filtered {len(filtered_df)} records from {len(df)} total.")

    # Reorder Columns
    preferred_order = [
        "Insurance Number", "Year", "Mail Number", "Mail Date",
        "Route", "Sector", "Flight Number", "Flight Date",
        "Complain Country", "Complain Location", "Commodity",
        "Weight KG", "Total Parcels", "Damaged Parcels",
        "Complain Type", "Other Complain",
        "Claimant Name", "Claimant Type",
        "Claim Date", "Insurance Sent",
        "Compensation Requested", "Req. Currency", "Requested in EGP", "Requested in USD", "Req. Date",
        "Compensation Accepted", "Acc. Currency", "Accepted in EGP", "Accepted in USD", "Acc. Date",
        "Status", "Other Status","Notes"
    ]

    existing_cols = [c for c in preferred_order if c in filtered_df.columns]
    remaining_cols = [c for c in filtered_df.columns if c not in existing_cols]
    filtered_df = filtered_df[existing_cols + remaining_cols]

    # Data Table
    st.subheader("Claims Table")
    st.dataframe(filtered_df, width="stretch")


    st.subheader("Download Table")


    def download_excel(df):
        output = io.BytesIO()

        # Write DataFrame (start at row 6)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name="Claims_Report",
                startrow=5  # leave space for logo & header
            )

        output.seek(0)

        wb = load_workbook(output)
        ws = wb["Claims_Report"]

        # Insert Logo
        logo = XLImage("Assets/EgyptAir_Cargo_Claims_Manager_Blue_logo.png")  # path or absolute path
        logo.width  = int(555 * 0.75)   # ≈ 1578
        logo.height = int(60 * 0.75)    # ≈ 165
        ws.add_image(logo, "A1")

        # Custom Header Text
        ws["A3"] = "Mail Claims Report"
        ws["A4"] = f"Generated on: {datetime.now().strftime('%d %b %Y %H:%M')}"

        ws["A3"].font = Font(bold=True, size=14)
        ws["A4"].font = Font(size=10)

        # Style Table Header Row
        header_row = 6  # because startrow=5

        header_fill = PatternFill(
            start_color="00265D",
            end_color="00265D",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None),
        )

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = no_border

        # Apply Filter
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A6:{last_col}{ws.max_row}"


        DATA_START_ROW  = header_row + 1  # row 7

        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(
            min_row=DATA_START_ROW,
            max_row=ws.max_row,
            min_col=1,
            max_col=len(df.columns)
        ):
            for cell in row:
                cell.alignment = left_align
                cell.border = thin_border


        # Auto Column Width
        HEADER_ROW = 6
        MAX_WIDTH = 40

        for i, col in enumerate(df.columns, 1):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, MAX_WIDTH)


        # Save back to memory
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        return final_output

    def excel_to_pdf(excel_bytes):

        # Save Excel to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_bytes.getvalue())
            excel_path = tmp_excel.name

        pdf_path = excel_path.replace(".xlsx", ".pdf")

        # Open Excel COM
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Worksheets[0]

        # Export as PDF
        ws.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF

        wb.Close(False)
        excel.Quit()

        # Read PDF back to memory
        with open(pdf_path, "rb") as f:
            pdf_bytes = io.BytesIO(f.read())

        # Cleanup
        os.remove(excel_path)
        os.remove(pdf_path)

        pdf_bytes.seek(0)
        return pdf_bytes

    # Generate Excel ONLY when clicked
    if "excel_file" not in st.session_state:
        st.session_state.excel_file_mail = None

    if "pdf_file" not in st.session_state:
        st.session_state.pdf_file_mail = None



    # To generate reports and download buttons
    if st.button("Generate Report"):
        #Generate excel
        st.session_state.excel_file_mail = download_excel(filtered_df)

        # Generate pdf
        st.session_state.pdf_file_mail = excel_to_pdf(st.session_state.excel_file_mail)


    #TimeStamp for file name
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    

    col1, col2, col3 = st.columns(3)

    with col1:
        if st.session_state.excel_file_mail:
            with stylable_container(
            key="excl_button",
            css_styles="""
                button {
                    background-color: #217346 !important;
                    color: white !important;
                    border-radius: 18px !important;
                    border: none !important;
                    transition: all 0.2s ease !important;
                }
                button:hover {
                    background-color: #f0f2f6 !important;
                    border-color: #000000 !important;
                    color: #000000 !important;
                    transform: translateY(-1px);
                    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.15) !important;
                }
            """,
        ):
                st.download_button(
                    "Download Excel",
                    data=st.session_state.excel_file_mail,
                    file_name=f"mail_claims_report_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=insert_log,
                    kwargs={"action": "Downloaded Excel file"}
                )

    with col2:
        if st.session_state.pdf_file_mail:
            with stylable_container(
            key="pdf_button",
            css_styles="""
                button {
                    background-color: #FF0000 !important;
                    color: white !important;
                    border-radius: 18px !important;
                    transition: all 0.2s ease !important;
                    border: none !important;
                }
                button:hover {
                    background-color: #f0f2f6 !important;
                    border-color: #000000 !important;
                    color: #000000 !important;
                    transform: translateY(-1px);
                    box-shadow: 0 4px 10px rgba(0, 0, 0, 0.15) !important;
                }
            """,
        ):
                st.download_button(
                    "Download PDF",
                    data=st.session_state.pdf_file_mail,
                    file_name=f"mail_claims_report_{timestamp}.pdf",
                    mime="application/pdf",
                    on_click=insert_log,
                    kwargs={"action": "Downloaded PDF file"}
                )


    # Static Charts
    st.subheader("Charts / Diagrams")

    country_count = filtered_df.groupby("Complain Country").size().reset_index(name="Count")
    st.plotly_chart(
        px.bar(country_count, x="Complain Country", y="Count", title="Claims by Country"),
        width="stretch"
    )

    # Claims Trend Over Time
    trend_df = (
            filtered_df
            .groupby("Complain Location")
            .size()
            .reset_index(name="Claims Count")
        )

    fig = px.bar(
            trend_df,
            x="Complain Location",
            y="Claims Count",
            title="Claims Trend Over Time"
        )
    st.plotly_chart(fig, width="stretch")


    # Analytics & Visualization
    st.subheader("Analytics & Visualizations")

    analysis_type = st.selectbox(
        "Select Analysis Type",
        [
            "Custom Chart Builder",
            "Top 10 by Category",
            "Damage Ratio Analysis"
        ]
    )

    # Custom Chart Builder
    if analysis_type == "Custom Chart Builder":
        chart_type = st.selectbox(
            "Chart Type",
            ["Bar", "Stacked Bar", "Pie", "Treemap", "Histogram"]
        )

        # Select X axis column (categorical or date)
        x_col = st.selectbox("X Axis", filtered_df.columns)

        # Identify numeric columns
        numeric_cols = filtered_df.select_dtypes(include="number").columns.tolist()

        # Y axis options depending on chart type
        y_col = None
        color_col = None  # for Stacked Bar and Treemap
        if chart_type in ["Bar", "Stacked Bar"] and numeric_cols:
            y_col = st.selectbox("Y Axis", numeric_cols)
            if chart_type == "Stacked Bar":
                # Suggest a categorical column for color
                categorical_cols = filtered_df.select_dtypes(
                    include=["object", "category"]
                ).columns.tolist()
                if categorical_cols:
                    color_col = st.selectbox("Stacked by (color)", categorical_cols)
        elif chart_type in ["Pie"] and numeric_cols:
            y_col = st.selectbox("Values", numeric_cols, index=0)
        elif chart_type == "Treemap":
            # Choose path (hierarchy)
            categorical_cols = filtered_df.select_dtypes(
                include=["object", "category"]
            ).columns.tolist()
            if len(categorical_cols) >= 2:
                path = st.multiselect("Treemap Path", categorical_cols, default=categorical_cols[:2])
            else:
                path = categorical_cols

        # Build Chart
        fig = None
        if chart_type == "Bar":
            fig = px.bar(filtered_df, x=x_col, y=y_col)
        elif chart_type == "Stacked Bar":
            fig = px.bar(filtered_df, x=x_col, y=y_col, color=color_col)
        elif chart_type == "Pie":
            fig = px.pie(filtered_df, names=x_col, values=y_col)
        elif chart_type == "Treemap":
            fig = px.treemap(filtered_df, path=path, values=y_col)
        elif chart_type == "Histogram":
            fig = px.histogram(filtered_df, x=x_col, color=color_col)

        # Show chart
        if fig:
            st.plotly_chart(fig, width="stretch")



    # Top 10 by Category
    elif analysis_type == "Top 10 by Category":
        cat_col = st.selectbox(
            "Select Category",
            ["Complain Country", "Complain Location", "Commodity", "Sector", "Claimant Type"]
        )

        top_df = (
            filtered_df
            .groupby(cat_col)
            .size()
            .reset_index(name="Claims")
            .sort_values("Claims", ascending=False)
            .head(10)
        )

        fig = px.bar(
            top_df,
            x=cat_col,
            y="Claims",
            text="Claims",
            title=f"Top 10 {cat_col} by Claims"
        )
        st.plotly_chart(fig, width="stretch")

    # Damage Ratio Analysis
    elif analysis_type == "Damage Ratio Analysis":
        df_ratio = filtered_df.copy()
        df_ratio["Damage Ratio"] = (
            df_ratio["Damaged Parcels"] / df_ratio["Total Parcels"]
        ).fillna(0)

        fig = px.bar(
            df_ratio,
            x="Commodity",
            y="Damage Ratio",
            title="Damage Ratio by Commodity"
        )
        st.plotly_chart(fig, width="stretch")


    # Aggregation Section
    st.subheader("Aggregation / Grouping")

    col1, col2, col3 = st.columns(3)

    with col1:
        group_col = st.selectbox(
            "Group By",
            options=filtered_df.columns
        )

    with col2:
        agg_func = st.selectbox(
            "Aggregation Function",
            ["count", "sum", "mean", "max", "min"]
        )

    with col3:
        numeric_cols = filtered_df.select_dtypes(include="number").columns.tolist()
        agg_col = st.selectbox(
            "Numeric Column",
            numeric_cols if agg_func != "count" else ["(count rows)"]
        )

    # Apply aggregation
    if agg_func == "count":
        agg_df = (
            filtered_df
            .groupby(group_col)
            .size()
            .reset_index(name="Count")
        )
    else:
        agg_df = (
            filtered_df
            .groupby(group_col)[agg_col]
            .agg(agg_func)
            .reset_index()
        )

    st.dataframe(agg_df, width="stretch")

